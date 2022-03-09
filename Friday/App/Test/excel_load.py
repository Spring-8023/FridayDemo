# -*- coding:utf-8 -*-
"""
Python 3.7
Author：Spring
Time：  2022/3/2 4:24 PM

"""
import openpyxl
import datetime

from Friday.App.Test.request_client import send


def read_excel(file_path,sheet_name):
    sheet_data = openpyxl.load_workbook(file_path)[sheet_name]
    lines_count = sheet_data.max_row
    cols_count = sheet_data.max_column
    # print(lines_count)
    # print(cols_count)
    cell_data = []
    for row in range(2, lines_count+1):
        # print
        lines = []

        for col in range(1, cols_count+1):
            cell_value = sheet_data.cell(row, col).value
            # cell_data["name"] = sheet_data.cell(row, col).value
            lines.append(cell_value)
        # print(lines)
            # print(cell_value)
        cell_data.append(lines)

    # print(cell_data)
    return cell_data

if __name__ == "__main__":
    data = read_excel('Friday_data.xlsx', 'Sheet1')
    host = "http://192.168.0.8"
    for resp in data:
        casename = resp[0]
        url = resp[1]
        method = resp[2]
        headers = resp[3]
        params = resp[4]

        print(casename)
        print(url)
        print(method)
        print(headers)
        print(params)
        print(type(params))
        # res = send(url=host+url, method=method, headers=headers, **params)
        # print(res.text)

